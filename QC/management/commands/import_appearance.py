import sys
import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from QC.models import AppearanceOption
from django.db import transaction

class Command(BaseCommand):
    help = (
        "Import Appearance Options from an Excel (.xlsx) file. "
        "Usage: python manage.py import_appearance /path/to/file.xlsx"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_path',
            type=str,
            help="Full filesystem path to the .xlsx file containing your appearance list."
        )

    def handle(self, *args, **options):
        excel_path = options['excel_path']

        # 1) Check that file exists
        if not os.path.isfile(excel_path):
            raise CommandError(f"File not found: {excel_path}")

        # 2) Try to open the workbook
        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Error opening Excel file: {e}")

        # 3) Assume the appearance names are in the very first worksheet, column A
        sheet = workbook.worksheets[0]

        # 4) Iterate down column A (skipping empty cells) and create AppearanceOption
        created_count = 0
        skipped_count = 0

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                cell_val = row[0]
                if cell_val is None:
                    continue

                name = str(cell_val).strip()
                if not name:
                    continue

                # If an AppearanceOption with this name already exists, skip
                obj, created = AppearanceOption.objects.get_or_create(name=name)
                if created:
                    created_count += 1
                else:
                    skipped_count += 1

        # 5) Print summary
        self.stdout.write(self.style.SUCCESS(
            f"Import finished. Created {created_count} new AppearanceOption(s), "
            f"skipped {skipped_count} already‐existing row(s)."
        ))
